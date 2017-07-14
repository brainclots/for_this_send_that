#!/usr/bin/env python

'''
Purpose:    Connect to a list of devices, stored in column 1 in a spreadsheet,
            then run the commands contained in column C. Column D should
            contain rollback commands. The rollback commands will be run only
            if the script is envoked with the '-r' option.
            | DeviceName | OS_Type  | Implementation_Cmds |   Rollback_Cmds  |
            | device1    | cisco_ios| commands to run     | rollback commands|
            | device2    | juniper  | commands to run     | rollback commands|
            |   etc...   | type     | commands to run     | rollback commands|
Author:
            ___  ____ _ ____ _  _    _  _ _    ____ ___ ___
            |__] |__/ | |__| |\ |    |_/  |    |  |  |    /
            |__] |  \ | |  | | \|    | \_ |___ |__|  |   /__
            Brian.Klotz@nike.com

Version:    1.1
Date:       July 2017
'''
import argparse
import netmiko
import getpass
import logging
import openpyxl

# Set up argument parser and help info
parser = argparse.ArgumentParser(
    formatter_class=argparse.RawDescriptionHelpFormatter,
    description='''\
    Connect to list of devices and run a set of commands on each.
    Format the Excel file as follows:
        | Column A   | Column B |     Column C        |     Column D     |
        +------------+----------+---------------------+------------------+
    Row1| DeviceName | OS_Type  | Implementation_Cmds |   Rollback_Cmds  |
    Row2| device1    | cisco_ios| commands to run     | rollback commands|
    Row3| device2    | juniper  | commands to run     | rollback commands|
    RowX|   etc...   | type     | commands to run     | rollback commands|
    (OS_Type can be either "juniper", "cisco_ios", "cisco_asa", or "cisco_ios_telnet")
    Optional: Include 'show' commands in Column E for on-screen
    verification of your implmentation commands (will be logged as well).

    Note: If the '-v' switch is given, you will be prompted to save changes,
    otherwise, changes are saved silently.
    ''')
always_required = parser.add_argument_group('always required')
always_required.add_argument("input_xlsx", nargs=1, help="Name of file containing \
                             devices and commands for each device",
                             metavar='<import_file>')
parser.add_argument('-v', '--verify',
                    help="Ask for verification before saving",
                    action="store_true")
parser.add_argument('-r', '--rollback', help="Run rollback commands",
                    action="store_true")

args = parser.parse_args()

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.FileHandler('output.log')
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)


def main():
    input_file = args.input_xlsx[0]
    input_info = open_file(input_file)

    username, password = get_creds()

    netmiko_exceptions = (netmiko.ssh_exception.NetMikoTimeoutException,
                          netmiko.ssh_exception.NetMikoAuthenticationException)

    logger.info('********* Starting run for %s devices **********',
                len(input_info))

    # Build dictionary of devices
    for row in range(1, len(input_info) + 1):
        device_dict = {'host': input_info[row]['host'],
                       'device_type': input_info[row]['device_type'],
                       'username': username,
                       'password': password,
                       'secret': password
                       }
        print('Connecting to ' + device_dict['host'] + ' (' +
              device_dict['device_type'] + ') ...')
        try:  # Connect to device
            connection = netmiko.ConnectHandler(**device_dict)
            logger.info('Successfully connected to %s', device_dict['host'])
            connection.enable()

            # Send commands
            if args.rollback:
                print('Sending rollback commands...')
                result = connection.send_config_set(
                                    input_info[row]['rollback_cmds'])
            else:
                print('Sending implementation commands...')
                if 'cisco' in device_dict['device_type']:
                    result = connection.send_config_set(
                                        input_info[row]['implementation_cmds'])
                elif 'juniper' in device_dict['device_type']:
                    result = connection.send_config_set(
                                        input_info[row]['implementation_cmds'],
                                        exit_config_mode=False)
            print('Finished sending commands.')
            indented_lines = indentem(result)
            logger.info('Actions: \n%s', indented_lines)

            # Run 'show' commands, if present
            if input_info[row]['verification_cmds']:
                verify_config(connection, input_info[row]['verification_cmds'],
                              input_info[row]['host'])

            # Determine whether to save
            if args.verify:
                yes_or_no = ask_to_save()
                if yes_or_no == 'y':
                    save_now(connection, device_dict['device_type'])
                    logger.info('Saved config changes on %s',
                                device_dict['host'])
                else:
                    print('Changes NOT saved!')
                    logger.info('Changes NOT saved, roll back or reboot')
            else:
                save_now(connection, device_dict['device_type'])
                logger.info('Saved changes on %s' % device_dict['host'])

            # Disconnect from device
            print('Disconnecting...')
            logger.info('Disconnecting from %s' % device_dict['host'])
            connection.disconnect()

        except netmiko_exceptions as e:
            print('Failed to connect: %s' % e)
            logger.error('Failed to connect %s', e)
    print('Completed. See "output.log" for results.')
    logger.info('********* End of run for %s devices **********',
                len(input_info))


def open_file(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    input_info = {}
    for row in range(2, ws.max_row + 1):
        device = ws['A' + str(row)].value
        # Subtract 1 from the row so that devices are numbered 1, 2, 3...
        input_info[row - 1] = {'host': device,
                               'device_type': ws['B' + str(row)].value,
                               'implementation_cmds': ws['C' + str(row)].value,
                               'rollback_cmds': ws['D' + str(row)].value,
                               'verification_cmds': ws['E' + str(row)].value
                               }
    return input_info


def get_creds():  # Prompt for credentials
    username = getpass.getuser()
#   username = raw_input('User ID: ')
    try:
        password = getpass.getpass()
        return username, password
    except KeyboardInterrupt:
        print('\n')
        exit()


def verify_config(connection, commands, hostname):
    proof = connection.send_command(commands)
    print('Output from Verification commands:'),
    print('\"' + commands + '\"')
    print(proof)
    outfile = hostname + '.show'
    f = open(outfile, 'w')
    f.write(hostname + '# ' + commands + '\n')
    f.write(proof)
    f.close()
    indented_lines = indentem(proof)
    logger.info('Verification commands: "%s" \n%s' %
                (commands, indented_lines))
    return


def ask_to_save():
    answer = raw_input('Save changes? (y/n) ')
    if answer.lower() == 'y':
        return answer


def save_now(connection, device_type):
    print('Saving config changes...')
    if 'cisco' in device_type:
        connection.send_command("write mem")
    elif device_type == 'juniper':
        connection.commit(and_quit=True)


def indentem(lines):
    lines = '    ' + lines
    lines = lines.replace('\n', '\n    ')
    return lines


main()
